from configparser import SafeConfigParser

import requests
import yaml
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.runtime.http.request_options import RequestOptions
from openpyxl import load_workbook

TESTBED_FILE_LOCAL_NAME = "./tmp/testbed_tracker_downloaded.xlsx"
COL_ROUTER_NAME = 1
COL_IP = 2
COL_PORT1 = 4
COL_PORT2 = 5
COL_TGEN = 6
COL_MH_CONFIG = 9
COL_PROJECT = 11
COL_RESERVATION_STATUS = 12


def download_testbed_excel_file():
    parser = SafeConfigParser(interpolation=None)
    parser.read("settings.cfg")

    username = parser.get("user_credentials", "username")
    password = parser.get("user_credentials", "password")
    testbed_file_url = parser.get("default", "testbed_file_url")

    ctx_auth = AuthenticationContext(testbed_file_url)
    ctx_auth.acquire_token_for_user(username, password)

    options = RequestOptions(testbed_file_url)
    ctx_auth.authenticate_request(options)

    req = requests.get(testbed_file_url,
                       headers=options.headers,
                       verify=True,
                       allow_redirects=True)
    output = open(TESTBED_FILE_LOCAL_NAME, "wb")
    output.write(req.content)
    output.close()


def open_excel_worksheet():
    wb = load_workbook(filename=TESTBED_FILE_LOCAL_NAME,
                       data_only=True,
                       keep_links=False)
    ws = wb.active
    return ws


def construct_device_name(ws, row):
    device = ws.cell(row=row, column=COL_ROUTER_NAME).value
    port2 = None
    port = None
    device_name = ""
    if device is None:
        device_name = "break"
    else:
        port = ws.cell(row=row, column=COL_PORT1).value
        if port is None:
            device_name = "continue"
        else:
            device_name = device + "_" + str(port)
            if ws.cell(row=row, column=COL_PORT2).value is not None:
                port2 = ws.cell(row=row, column=COL_PORT2).value
                # device_name += "_"
                # device_name += str(port2)
    return device_name, port, port2


def get_reservation_status(ws, row):
    return ws.cell(row=row, column=COL_RESERVATION_STATUS).value


def construct_testbed_yaml_dict_from_excel_ws(ws):
    """
    Construct a dict from the excel worksheet object. This function assumes the
    following format of excel sheet.
    Col A: Router name
    Col B: Router's telnet IP address
    Col D: Telnet console port 1
    Col E: Telnet console port 2
    Col F: TGEN connections to the router
    Col I: MH Config
    Col K: Project name

    :param ws:
        Excel worksheet object.

    :return:
        Yaml dictionary
    """

    yaml_dict = {"devices": {}}
    for row in range(2, 1000):
        device_name, port, port2 = construct_device_name(ws, row)
        if device_name == "break":
            break
        elif device_name == "continue":
            continue
        ip = ws.cell(row=row, column=COL_IP).value
        if ip is None:
            continue
        if port2 is None:
            device_dict = {
                device_name: {
                    "connections": {
                        "default": {
                            "class": "unicon.Unicon"
                        },
                        "a": {
                            "ip": ip,
                            "port": port,
                            "protocol": "telnet"
                        },
                    },
                    "credentials": {
                        "default": {
                            "password": "lab123",
                            "username": "lab"
                        },
                        "alternate0": {
                            "password": "lab123",
                            "username": "root"
                        },
                        "alternate1": {
                            "password": "test123",
                            "username": "root"
                        },
                        "alternate2": {
                            "password": "root!lab",
                            "username": "root"
                        },
                        "alternate3": {
                            "password": "root123",
                            "username": "root1"
                        },
                        "enable": {
                            "password": "lab123"
                        },
                    },
                    "os": "iosxr",
                    "type": "router",
                    "custom": {
                        "execute_timeout": 80,
                        "configure_timeout": 65,
                        "abstraction": {
                            "order": ["os"]
                        },
                        "mh_config": ws.cell(row=row,
                                             column=COL_MH_CONFIG).value,
                        "tgen": ws.cell(row=row, column=COL_TGEN).value,
                        "project": ws.cell(row=row, column=COL_PROJECT).value,
                    },
                }
            }
        else:
            device_dict = {
                device_name: {
                    "connections": {
                        "default": {
                            "class": "unicon.Unicon"
                        },
                        "a": {
                            "ip": ip,
                            "port": port,
                            "protocol": "telnet"
                        },
                        "b": {
                            "ip": ip,
                            "port": port2,
                            "protocol": "telnet"
                        },
                    },
                    "credentials": {
                        "default": {
                            "password": "lab123",
                            "username": "lab"
                        },
                        "alternate0": {
                            "password": "lab123",
                            "username": "root"
                        },
                        "alternate1": {
                            "password": "test123",
                            "username": "root"
                        },
                        "alternate2": {
                            "password": "root!lab",
                            "username": "root"
                        },
                        "alternate3": {
                            "password": "root123",
                            "username": "root1"
                        },
                        "enable": {
                            "password": "lab123"
                        },
                    },
                    "os": "iosxr",
                    "type": "router",
                    "custom": {
                        "execute_timeout": 80,
                        "configure_timeout": 65,
                        "abstraction": {
                            "order": ["os"]
                        },
                        "mh_config": ws.cell(row=row,
                                             column=COL_MH_CONFIG).value,
                        "tgen": ws.cell(row=row, column=COL_TGEN).value,
                        "project": ws.cell(row=row, column=COL_PROJECT).value,
                    },
                }
            }
        yaml_dict["devices"][device_name] = device_dict[device_name]
    return yaml_dict


def write_yaml_dict_to_yaml_file(source_yaml_dict, dest_filename):
    with open(dest_filename, "w") as outfile:
        yaml.dump(source_yaml_dict, outfile, default_flow_style=False)